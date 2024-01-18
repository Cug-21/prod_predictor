from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pandas as pd
import finnhub
import matplotlib.pyplot as plt
from flask import request
import yfinance as yf

finnhub_client = finnhub.Client(api_key="cicrv2hr01qopecuh4o0cicrv2hr01qopecuh4og")

def index_to_column_letter(index):
    return get_column_letter(index)


def get_financial_ratios(stock):
    stats = stock.info
    try:

        current_ratio = stats['currentRatio']
        return_on_equity = stats['returnOnEquity']
        profit_margin = stats['profitMargins']
        price_to_book = stats['priceToBook']
        debt_to_equity = stats['debtToEquity']
        earnings_per_share = stats['trailingEps']
        forward_eps = stats['forwardEps']
        operating_margin = stats['operatingMargins']
        beta = stats['beta']
        quick_ratio = stats['quickRatio']
        peg_ratio = stats['pegRatio']
        short_ratio = stats['shortRatio']
        ebitda_margin = stats['ebitdaMargins']
        enterprise_value = stats['enterpriseValue']
        enterprise_to_ebitda = enterprise_value/stats['ebitda']

        ratios = {
            "Current Ratio": current_ratio,
            "Return on Equity": return_on_equity,
            "Profit Margin": profit_margin,
            "Price to Book": price_to_book,
            "Debt to Equity": debt_to_equity,
            "EPS": earnings_per_share,
            "Forward EPS": forward_eps,
            "Operating margin": operating_margin,
            "Beta": beta,
            "Quick Ratio": quick_ratio,
            "Peg Ratio": peg_ratio,
            "Short Ratio": short_ratio,
            "Ebitda Margin": ebitda_margin,
            "Enterprise Value": enterprise_value,
            "Enterprise to Ebitda": enterprise_to_ebitda
        }

        return ratios
    except KeyError as e:
        print(f"Key error: {e}")
        return {}

def financialRatiosReport(ws, stock):
    ratios = get_financial_ratios(stock)
    data_col = 1
    data_row = 1

    ws.cell(row=data_row, column=data_col, value="Financial Ratios")
    data_row += 1

    for ratio, value in ratios.items():
        ws.cell(row=data_row, column=data_col, value=ratio)
        ws.cell(row=data_row, column=data_col + 1, value=value)
        data_row += 1

        
def CompanyHolderReport(ws, stock):
    major_holders = stock.major_holders
    inst_holders = stock.institutional_holders
    fund_holders = stock.mutualfund_holders
    earnings = stock.earnings_dates
    print(major_holders)
    print(inst_holders)
    print(fund_holders)

    data_col = 1  
    data_row = 1 

    ws.cell(row=data_row, column=data_col, value="Major Holders Breakdown")
    data_row += 1

    for _, row in major_holders.iterrows():
        ws.cell(row=data_row, column=data_col, value=row.name)  # Description
        ws.cell(row=data_row, column=data_col + 1, value=row[0])  # Value
        data_row += 1

    data_row += 2 # Skip a row for separation

    data_row += 1


    ws.cell(row=data_row, column=data_col, value="Institutional Holders")
    data_row += 1
    ws.cell(row=data_row, column=data_col, value="Holder")
    ws.cell(row=data_row, column=data_col + 1, value="Shares")
    ws.cell(row=data_row, column=data_col + 2, value="% Held")
    ws.cell(row=data_row, column=data_col + 3, value="Value")
    data_row += 1
    for _, row in inst_holders.iterrows():
        ws.cell(row=data_row, column=data_col, value=row['Holder'])
        ws.cell(row=data_row, column=data_col + 1, value=row['Shares'])
        ws.cell(row=data_row, column=data_col + 2, value=row['pctHeld'])
        ws.cell(row=data_row, column=data_col + 3, value=row['Value'])
        data_row += 1

    # Mutual Fund Holders
    data_row += 1
    ws.cell(row=data_row, column=data_col, value="Mutual Fund Holders")
    data_row += 1
    ws.cell(row=data_row, column=data_col, value="Holder")
    ws.cell(row=data_row, column=data_col + 1, value="Shares")
    ws.cell(row=data_row, column=data_col + 2, value="% Held")
    ws.cell(row=data_row, column=data_col + 3, value="Value")
    data_row += 1
    for _, row in fund_holders.iterrows():
        ws.cell(row=data_row, column=data_col, value=row['Holder'])
        ws.cell(row=data_row, column=data_col + 1, value=row['Shares'])
        ws.cell(row=data_row, column=data_col + 2, value=row['pctHeld'])
        ws.cell(row=data_row, column=data_col + 3, value=row['Value'])
        data_row += 1


def get_financial_data(stock):
    df = stock.cashflow
    print(df.columns)
    if 'Free Cash Flow' in stock.cashflow.index:
        free_cash_flow_row = stock.cashflow.loc['Free Cash Flow']
        free_cash_flow_row = pd.to_numeric(free_cash_flow_row, errors='coerce')
        average_free_cash_flow = free_cash_flow_row.mean(skipna=True)
    else:
        average_free_cash_flow = 0

    financial_data = {
        'FreeCashFlow': average_free_cash_flow,
    }
    
    print(financial_data)
    return financial_data

def dcfReport(ws, stock, dcf_params):
    growth_rate = dcf_params.get('growth_rate', 3) / 100
    discount_rate = dcf_params.get('discount_rate', 5) / 100
    terminal_growth_rate = dcf_params.get('terminal_growth_rate', 2) / 100
    years = dcf_params.get('years', 5)  

    financial_data = get_financial_data(stock)  

    dcf_value, discounted_cash_flows = calculate_dcf(stock, financial_data, growth_rate, discount_rate, terminal_growth_rate, years)

    row = 1
    for i, value in enumerate(discounted_cash_flows, start=1):
        ws.cell(row=row, column=1, value=f"Year {i}")
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws.cell(row=row, column=1, value="Total DCF Value")
    ws.cell(row=row, column=2, value=dcf_value)


def calculate_dcf(stock, financial_data, growth_rate, discount_rate, terminal_growth_rate, years):
    future_cash_flows = []
    years = int(years)
    for i in range(1, years + 1):
        future_cash_flow = financial_data['FreeCashFlow'] * ((1 + growth_rate) ** i)
        future_cash_flows.append(future_cash_flow)
    # Discount future cash flows back to present value
    discounted_cash_flows = [cf / ((1 + discount_rate) ** i) for i, cf in enumerate(future_cash_flows, 1)]

    # Calculate terminal value
    terminal_value = future_cash_flows[-1] * (1 + terminal_growth_rate) / (discount_rate - terminal_growth_rate)
    terminal_value_discounted = terminal_value / ((1 + discount_rate) ** years)

    # DCF valuation
    dcf_value = sum(discounted_cash_flows) + terminal_value_discounted

    return dcf_value, discounted_cash_flows

def create_dcf_chart(discounted_cash_flows):
    years = range(1, len(discounted_cash_flows) + 1)
    plt.bar(years, discounted_cash_flows)
    plt.xlabel('Year')
    plt.ylabel('Discounted Cash Flow')
    plt.title('Discounted Cash Flow Over Time')

    # Save the plot as an image
    plt.savefig('/path/to/dcf_chart.png')
    plt.close()

def insert_chart_into_excel(filename, chart_filename):
    workbook = load_workbook(filename)

    # Select the 'DCF Model' sheet
    if "DCF Model" in workbook.sheetnames:
        sheet = workbook["DCF Model"]
    else:
        print("DCF Model sheet not found")
        return
    img = Image(chart_filename)
    # Specify the cell for the top left corner of the image
    sheet.add_image(img, 'A10')  
    workbook.save(filename) 

def companyQuoteReport(ws, stock):
    data = request.get_json()
    ticker = data.get('symbol')
   
    quote = {}
    recommendation_trends = []
    if ticker:
        recommendation_trends = finnhub_client.recommendation_trends(ticker)
        quote = finnhub_client.quote(symbol=ticker)

    recommendation_col = 1
    recommendation_row = 1  # Adjust the starting row as necessary, depending on the previous data
    ws.cell(row=recommendation_row, column=recommendation_col, value="Recommendation Trends")
    recommendation_row += 1

    for trend in recommendation_trends:
        ws.cell(row=recommendation_row, column=recommendation_col, value="Period")
        ws.cell(row=recommendation_row, column=recommendation_col + 1, value=trend["period"])
        recommendation_row += 1

        ws.cell(row=recommendation_row, column=recommendation_col, value="Strong Buy")
        ws.cell(row=recommendation_row, column=recommendation_col + 1, value=trend["strongBuy"])
        recommendation_row += 1

        ws.cell(row=recommendation_row, column=recommendation_col, value="Buy")
        ws.cell(row=recommendation_row, column=recommendation_col + 1, value=trend["buy"])
        recommendation_row += 1

        ws.cell(row=recommendation_row, column=recommendation_col, value="Hold")
        ws.cell(row=recommendation_row, column=recommendation_col + 1, value=trend["hold"])
        recommendation_row += 1

        ws.cell(row=recommendation_row, column=recommendation_col, value="Sell")
        ws.cell(row=recommendation_row, column=recommendation_col + 1, value=trend["sell"])
        recommendation_row += 1

        ws.cell(row=recommendation_row, column=recommendation_col, value="Strong Sell")
        ws.cell(row=recommendation_row, column=recommendation_col + 1, value=trend["strongSell"])
        recommendation_row += 1

        recommendation_row += 1 
 
    
def balanceSheetReport(ws, stock):
    balance_sheet = stock.balance_sheet
    start_row = ws.max_row + 2 if ws.max_row > 1 else 1
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.value = 'Balance Sheet'
    title_cell.font = Font(name='Calibri', bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    for col_num, column_title in enumerate(balance_sheet.columns, start=1):
        header_cell = ws.cell(row=start_row + 1, column=col_num)
        header_cell.value = str(column_title)  # Convert to string to avoid datetime format

        header_cell.font = Font(name='Calibri', bold=True, size=14)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, (index, row_data) in enumerate(balance_sheet.iterrows(), start=start_row + 2):
        ws.cell(row=row_num, column=1, value=str(index))  # Writing the index (e.g., line item name)
        for col_num, value in enumerate(row_data, start=2):  # Start from the second column
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value if pd.notna(value) else ''  # Check for NaN values
            # Apply data styles
            cell.font = Font(name='Calibri', size=12)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))


def cashflowReport(ws, stock):
    cashflow = stock.cashflow
   
    start_row = ws.max_row + 2 if ws.max_row > 1 else 1

    # Adding title for the Cashflow Statement
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.value = 'Cashflow Statement'
    title_cell.font = Font(name='Calibri', bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Writing the header (column titles)
    for col_num, column_title in enumerate(cashflow.columns, start=2):  # Start from the second column
        header_cell = ws.cell(row=start_row + 1, column=col_num)
        header_cell.value = str(column_title)  # Convert to string to avoid datetime format
        # Apply header styles
        header_cell.font = Font(name='Calibri', bold=True, size=14)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Writing the data including row labels
    for row_num, (index, row_data) in enumerate(cashflow.iterrows(), start=start_row + 2):
        # Writing the row label
        label_cell = ws.cell(row=row_num, column=1)
        label_cell.value = str(index)  # Row label (e.g., 'Free Cash Flow')
        label_cell.font = Font(name='Calibri', size=12)

        # Writing the row data
        for col_num, value in enumerate(row_data, start=2):  # Start from the second column
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value if pd.notna(value) else ''  # Check for NaN values
            # Apply data styles
            cell.font = Font(name='Calibri', size=12)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

def incomeStatementReport(ws, stock):
    income_stmt = stock.income_stmt

    # Start row for the Income Statement title and data
    start_row = ws.max_row + 2 if ws.max_row > 1 else 1

    # Adding title for the Income Statement
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.value = 'Income Statement'
    title_cell.font = Font(name='Calibri', bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # Writing the header (column titles)
    for col_num, column_title in enumerate(income_stmt.columns, start=1):
        header_cell = ws.cell(row=start_row + 1, column=col_num)
        header_cell.value = str(column_title)  # Convert to string to avoid datetime format
        # Apply header styles
        header_cell.font = Font(name='Calibri', bold=True, size=14)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Writing the data
    for row_num, (index, row_data) in enumerate(income_stmt.iterrows(), start=start_row + 2):
        ws.cell(row=row_num, column=1, value=str(index))  # Writing the index (e.g., line item name)
        for col_num, value in enumerate(row_data, start=2):  # Start from the second column
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value if pd.notna(value) else ''  # Check for NaN values
            # Apply data styles
            cell.font = Font(name='Calibri', size=12)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def dividendAnalysis(ws, stock):
    dividends = stock.dividends
    data_col = 1
    data_row = 1

    ws.cell(row=data_row, column=data_col, value="Dividend History")
    data_row += 1

    for date, dividend in dividends.items():
        ws.cell(row=data_row, column=data_col, value=str(date))
        ws.cell(row=data_row, column=data_col + 1, value=dividend)
        data_row += 1

