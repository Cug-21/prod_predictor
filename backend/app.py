from flask import Flask, jsonify, request, send_file
import yfinance as yf
from flask_cors import CORS
from datetime import datetime
import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
import requests
import openpyxl
import traceback 
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from eco_market_analysis import get_market_analysis, get_eco_analysis
from market_model import load_market_data_news, get_market_health_score
import chat_model
import ai_model
from excel_report import incomeStatementReport, financialRatiosReport, balanceSheetReport, companyQuoteReport, cashflowReport, dcfReport, CompanyHolderReport
from eco_model import eco_score, economic_features


app = Flask(__name__)

CORS(app, resources={r"/*": {"origins": "*", "supports_credentials": True}}, 
     allow_headers=[
        "Content-Type", "Authorization", "Access-Control-Allow-Credentials"
     ],
     supports_credentials=True)



@app.route('/stock/<tickers>', methods=['GET'])
def get_stock_data(tickers):
    tickers = tickers.split(',')
    stock_data_list = []

    for ticker in tickers:
        stock = yf.Ticker(ticker)
        data = stock.history(period="2d")

        if not data.empty:
            current_price = data['Close'].iloc[-1]
            previous_price = data['Close'].iloc[-2]
            percentage_change = ((current_price - previous_price) / previous_price) * 100
            now = datetime.now()
            market_open_time = datetime(now.year, now.month, now.day, 9, 30)  # 9:30 AM
            market_close_time = datetime(now.year, now.month, now.day, 16, 0)  # 4:00 PM
            market_status = "open" if market_open_time <= now <= market_close_time else "closed"
            stock_data_list.append({
                'symbol': ticker,
                'price': current_price,
                'previousPrice': previous_price,
                'percentageChange': percentage_change,
                'marketStatus': market_status
            })

    return jsonify(stock_data_list)


@app.route('/GovTrades', methods=['POST'])
def GovTrades():
    data = request.json
    name = data.get('name', '')
    Ticker = data.get('Ticker', '')
    trades = []
    representatives = []
    party_counts = {'D': 0, 'R': 0, 'I': 0, ' ': 0}
    transaction_count = {'Purchase': 0, 'Sale': 0, 'Exercise': 0, 'Exchange': 0, 'Sale (Full)': 0}
    url = "https://api.quiverquant.com/beta/bulk/congresstrading"

    headers = {
        'Accept': "application/json",
        'Authorization': "Bearer d83982dc8d7fa9cfd45ee7cae570b3ddc8349f09"
    }
    query_params = {}
    if Ticker:
        query_params["ticker"] = Ticker

    elif name:
        query_params["name"] = name

    response = requests.get(url, headers=headers, params=query_params)
    if response.status_code == 200:
        all_trades = response.json()

        if Ticker:
            trades = [trade for trade in all_trades if trade['Ticker'] == Ticker]
        elif name:
            trades = [trade for trade in all_trades if trade['Representative'].lower() == name.lower()]

        representatives = sorted(list(set([trade['Representative'] for trade in all_trades])))
        for trade in trades:
            party_counts[trade['Party']] += 1
            transaction_count[trade['Transaction']] = transaction_count.get(trade['Transaction'], 0) + 1

    return jsonify({
        'name': name,
        'Ticker': Ticker,
        'trades': trades,
        'representatives': representatives,
        'party_counts': party_counts,
        'transaction_count': transaction_count
    })


@app.route('/aiChat', methods=['POST'])
def ai_chat():
    data = request.json
    ticker = data['ticker']
    combined_data = ai_model.main(ticker)
    chat_response = chat_model.get_chat_response(combined_data)

    return jsonify(chat_response)


@app.route('/userReport', methods=['POST', 'GET'])
def userReport():
    data = request.get_json()
    symbol = data.get('symbol')
    include_holders = data.get('include_holders', False)
    include_quote = data.get('include_quote', False)
    include_cashflow = data.get('include_cashflow', False)
    include_balancesheet = data.get('include_balancesheet', False)
    include_incomestatement = data.get('include_incomestatement', False)
    include_dcf = data.get('include_dcf', False)
    dcf_params = data.get('dcf_params', {})
    include_financial_ratios = data.get('include_financial_ratios', False)

    wb = openpyxl.Workbook()
    stock = yf.Ticker(symbol)
    
    if include_holders:
        ws_holders = wb.create_sheet(title="Holders Data")
        CompanyHolderReport(ws_holders, stock)

    if include_quote:
        ws_quote = wb.create_sheet(title="Quote Data")
        companyQuoteReport(ws_quote, stock)

    if include_cashflow or include_balancesheet or include_incomestatement or include_financial_ratios:
        ws_financials = wb.create_sheet(title="Financials Data")

    if include_cashflow:
        cashflowReport(ws_financials, stock)
        next_row = ws_financials.max_row + 3

    if include_balancesheet:
        balanceSheetReport(ws_financials, stock)
        next_row = ws_financials.max_row + 3

    if include_incomestatement:
        incomeStatementReport(ws_financials, stock)
        next_row = ws_financials.max_row + 3

    if include_dcf:
        ws_dcf = wb.create_sheet(title='DCF Model')
        dcfReport(ws_dcf, stock, dcf_params)

    if include_financial_ratios:
        ws_financial_ratios = wb.create_sheet(title="Financial Ratios")
        financialRatiosReport(ws_financial_ratios, stock)

    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    filename = f"{symbol}_ExcelReport.xlsx"
    return send_file(filename, as_attachment=True)



market_cache = {
    'market_health_score': None,
    'market_health_data': None,
    'market_last_updated': None
}

@app.route('/market_health', methods=['GET'])
def market_health():
    try:
        current_time = datetime.now()
        last_updated = market_cache.get('market_last_updated')
        if last_updated and (current_time.date() == last_updated.date()):
            score = market_cache['market_health_score']
        else:
            # If the cache is not valid, recalculate the score
            combined_data = load_market_data_news()
            market_cache['market_health_data']= combined_data,
            score = get_market_health_score(combined_data)
            market_cache['market_health_score'] = score
            market_cache['market_last_updated'] = current_time

        return jsonify({'market_health_score': score})
    except Exception as e:
        print(e)
        return jsonify({'error': str(e)}), 500

eco_cache = {
    'eco_health_score': None,
    'eco_health_data': None,
    'eco_last_updated': None
}


unused= ['unemployment_rate_annual_change', 'inflation_rate_data_annual_change', 
                  'cpi_data_annual_change', 'interest_rate_data_annual_change', 
                  'ppi_data_annual_change', 'bank_lending_data_annual_change', 
                  'balance_of_trade_data_annual_change', 'wage_growth_data_annual_change', 
                  'retail_sales_data_annual_change', 'housing_market_data_annual_change', 
                  'government_debt_data_annual_change', 'currency_exchange_annual_change', 
                  'bond_yield_annual_change']

@app.route('/eco_health', methods=['GET'])
def eco_health():
    try:
        current_time = datetime.now()
        last_updated = eco_cache.get('eco_last_updated')
        if last_updated and (current_time.date() == last_updated.date()):
            score = eco_cache['eco_health_score']
        else:
            combined_data = economic_features
            score = eco_score()  
            eco_cache['eco_health_score'] = score
            eco_cache['eco_health_data'] = combined_data
            eco_cache['eco_last_updated'] = current_time

        return jsonify({'eco_health_score': score})
    except Exception as e:
        print(e)
        return jsonify({'error': str(e)}), 500



    
def process_eco_data(eco_combined_data):
    important_keys = [
        'unemployment_rate_annual_change', 
        'inflation_rate_data_annual_change', 
        'cpi_data_annual_change', 
        'interest_rate_data_annual_change', 
    ]
    filtered_eco_combined_data = {}
    for key in important_keys:
        data_list = eco_combined_data.get(key, [])
        filtered_data = [
            data for data in data_list 
            if datetime.strptime(data['date'], '%Y-%m-%d %H:%M:%S').year >= 2020
        ]
        filtered_eco_combined_data[key] = filtered_data

    for key in filtered_eco_combined_data:
        df = pd.DataFrame(filtered_eco_combined_data[key])
        df['date'] = pd.to_datetime(df['date'])
        df.set_index('date', inplace=True)
        monthly_df = df.resample('Y').mean()
        monthly_df.reset_index(inplace=True)
        monthly_df['date'] = monthly_df['date'].dt.strftime('%Y-%m-%d')
        filtered_eco_combined_data[key] = monthly_df.reset_index().to_dict(orient='records')

    return filtered_eco_combined_data
    


analysis_cache = {
    'eco_analysis_score': None,
    'eco_analysis': None,
    'eco_last_updated_analysis': None,
    'market_analysis_score': None,
    'market_analysis': None,
    'market_last_updated_analysis': None,
}
    
@app.route('/eco_market_analysis', methods=['GET'])
def eco_market_analysis():
    print("eco_market_analysis route called")
    print("eco_market_analysis route called")
    try:
        current_time = datetime.now()

        # Fetch market and economic scores and data from respective caches
        market_score = market_cache['market_health_score']
        combined_data_market = market_cache['market_health_data']
        eco_score = eco_cache['eco_health_score']
        eco_combined_data = eco_cache['eco_health_data']

        # Check and update market analysis in analysis_cache if necessary
        if not is_cache_valid(analysis_cache['market_last_updated_analysis']):
            # Assuming get_market_analysis() processes the market data and returns the analysis
            market_analysis = get_market_analysis(market_score, combined_data_market)
            analysis_cache['market_analysis'] = market_analysis
            analysis_cache['market_analysis_score'] = market_score  # Or any other scoring logic you might have
            analysis_cache['market_last_updated_analysis'] = current_time
        else:
            # Use cached market analysis data
            market_analysis = analysis_cache['market_analysis']

        # Check and update economic analysis in analysis_cache if necessary
        if not is_cache_valid(analysis_cache['eco_last_updated_analysis']):
            # Process eco_combined_data
            processed_eco_data = process_eco_data(eco_combined_data)
            eco_analysis = get_eco_analysis(eco_score, processed_eco_data)
            analysis_cache['eco_analysis'] = eco_analysis
            analysis_cache['eco_analysis_score'] = eco_score 
            analysis_cache['eco_last_updated_analysis'] = current_time
        else:
            # Use cached economic analysis data
            eco_analysis = analysis_cache['eco_analysis']

        return jsonify({'market_health_analysis': market_analysis, 'eco_health_analysis': eco_analysis})

    except Exception as e:
        print(f"Exception in eco_market_analysis: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def is_cache_valid(last_updated):
    current_time = datetime.now()
    return last_updated and (current_time.date() == last_updated.date())



if __name__ == '__main__':
    app.run(debug=True)